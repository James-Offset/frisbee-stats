"""A sub-programme to extract the raw data from the excel file for use in the main code doc"""

import os
import openpyxl

class DataExtractor():
    """A class to hold the functions that get the data out of the excel file"""

    def __init__(self):
        pass
 
    def import_stock_data(self, filename):
        """Imports the data we have from the excel file that tells us what happened in the tournament"""

        # open the excel doc
        self.workbook = openpyxl.load_workbook(filename, data_only=True)

        # extract relevant data points
        self.set_up_tournament_metadata()
        self.define_the_team_roster()
        self.all_games_data_check()

        return self.tournament_metadata, self.roster, self.raw_game_data

    
    def set_up_tournament_metadata(self):
        """Captures key tournament metadata"""

        # open the front page
        sheet = self.workbook["Front Page"]

        # set up the dictionary and capture key info
        self.tournament_metadata = {
            "Team Name" : sheet['B1'].value,
            "Tournament Name" : sheet['B2'].value,
            "Players per Point" : sheet['C4'].value,
            "Environment" : sheet['B3'].value,
        }


    def define_the_team_roster(self):
        """Looks at the chosen excel file and identifies the team roster"""

        # set up a dictionary for the roster
        self.roster = {}

        sheet = self.workbook["Front Page"]

        for row in range(8, 40):
            # capture player name and number
            player_name = sheet.cell(row=row, column=2).value
            player_number = sheet.cell(row=row, column=3).value

            if player_name == None:
                break
            else:
                self.roster[player_name] = {}
                self.roster[player_name]["Player Number"] = player_number

    def all_games_data_check(self):
        """Checks which sheets have data and then calls the extraction function for those"""

        # set up a dictionary to store game data
        self.raw_game_data = {}

        # keep a tally of how many games were played
        game_number = 0

        # get the sheetnames
        data_sheets = self.workbook.sheetnames

        # go through each sheet and check for data
        for sheet_name in data_sheets:

            # sheet name needs to be "Game(x)"
            if sheet_name[:4] == "Game":
                sheet_to_review = self.workbook[sheet_name]

                # look at a cell that should have data
                test_point = sheet_to_review['G4'].value

                try:
                    int_test_point = int(test_point)
                except Exception:
                    break
                else:
                    game_number += 1
                    game_number_string = str(game_number)
                    self.extract_game_data(sheet_to_review, game_number_string)

    
    def extract_game_data(self, sheet, game_number_string):
        """Goes into the specified sheet and extracts the raw data needed for later analysis"""

        game_opponent = sheet["B1"].value
        team_start_on_defence = sheet['B5'].value

        # set up the reference name of the game and log info
        game_name = "Game " + game_number_string
        self.raw_game_data[game_name] = {
            "Opponent" : game_opponent,
            "Starting on Defence" : team_start_on_defence,
            "Turns per Point" : [],
            "Active Players" : [],
        }

        # add an entry for each player to capture whether they were playing that point
        for player in self.roster:
            self.raw_game_data[game_name][player] = []

        # loop through each row and capture the data
        for row in range(2,31):
            number_of_turns = sheet.cell(row=row, column=7).value

            # check there is data here
            try:
                turns_integer = int(number_of_turns)
            except Exception:
                break
            else:
                # log the number of turns
                self.raw_game_data[game_name]["Turns per Point"].append(turns_integer)

                # for each player, log if they were on that point
                column = 8
                active_players = []
                for player in self.roster:
                    player_presence = sheet.cell(row=row, column=column).value

                    if player_presence == 1:
                        # player was on the pitch
                        self.raw_game_data[game_name][player].append(1)
                        active_players.append(player)
                    else:
                        # player was not on pitch
                        self.raw_game_data[game_name][player].append(0)
                    column += 1
                
                # update the raw data dict with the list of active players
                self.raw_game_data[game_name]["Active Players"].append(active_players)

